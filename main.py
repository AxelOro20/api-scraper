"""
main.py  ─  API REST del Scraper de Precios
Uso:
    uvicorn main:app --reload

Documentación interactiva:
    http://localhost:8000/docs      ← Swagger UI
    http://localhost:8000/redoc     ← ReDoc
"""

import io
import statistics
from datetime import timedelta
from typing import Optional

import pandas as pd
from fastapi import FastAPI, Depends, HTTPException, status, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
import schemas
from auth import (
    hash_password, verificar_password,
    crear_token, get_usuario_actual
)
from database import engine, get_db
from scraper import hacer_scraping, TIENDAS_CONFIG

# Crear tablas al iniciar
models.Base.metadata.create_all(bind=engine)

# ── App ────────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="🛒 Scraper de Precios API",
    description="""
API REST para comparar precios entre tiendas en línea.

## Flujo de uso
1. **Registrarse** en `/auth/registro`
2. **Iniciar sesión** en `/auth/login` → obtener token JWT
3. Usar el token en el botón **Authorize 🔒** de arriba
4. **Buscar productos** en `/scraper/buscar`
5. **Comparar** precios entre tiendas en `/comparador/comparar`
6. **Exportar** resultados en `/exportar/csv` o `/exportar/excel`
    """,
    version="1.0.0",
    contact={"name": "Scraper de Precios", "email": "contacto@ejemplo.com"},
)


# ══════════════════════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.post(
    "/auth/registro",
    response_model=schemas.UsuarioRespuesta,
    status_code=status.HTTP_201_CREATED,
    tags=["🔐 Autenticación"],
    summary="Registrar nuevo usuario",
)
def registro(datos: schemas.UsuarioRegistro, db: Session = Depends(get_db)):
    """Crea una cuenta nueva. El email debe ser único."""
    if db.query(models.Usuario).filter(models.Usuario.email == datos.email).first():
        raise HTTPException(status_code=400, detail="El email ya está registrado")

    usuario = models.Usuario(
        nombre   = datos.nombre,
        email    = datos.email,
        password = hash_password(datos.password),
    )
    db.add(usuario)
    db.commit()
    db.refresh(usuario)
    return usuario


@app.post(
    "/auth/login",
    response_model=schemas.Token,
    tags=["🔐 Autenticación"],
    summary="Iniciar sesión",
)
def login(datos: schemas.UsuarioLogin, db: Session = Depends(get_db)):
    """Devuelve un token JWT válido por 24 horas."""
    usuario = db.query(models.Usuario).filter(models.Usuario.email == datos.email).first()
    if not usuario or not verificar_password(datos.password, usuario.password):
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")

    token = crear_token({"sub": usuario.email}, timedelta(hours=24))
    return {"access_token": token, "token_type": "bearer"}


@app.get(
    "/auth/yo",
    response_model=schemas.UsuarioRespuesta,
    tags=["🔐 Autenticación"],
    summary="Ver mi perfil",
)
def mi_perfil(usuario: models.Usuario = Depends(get_usuario_actual)):
    """Devuelve los datos del usuario autenticado."""
    return usuario


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

@app.get(
    "/scraper/tiendas",
    tags=["🤖 Scraper"],
    summary="Ver tiendas disponibles",
)
def listar_tiendas():
    """Lista las tiendas que se pueden scrapear."""
    return {"tiendas": list(TIENDAS_CONFIG.keys())}


@app.post(
    "/scraper/buscar",
    response_model=schemas.BusquedaRespuesta,
    status_code=status.HTTP_201_CREATED,
    tags=["🤖 Scraper"],
    summary="Buscar producto en una tienda",
)
def buscar_producto(
    datos: schemas.BusquedaRequest,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """
    Inicia el scraping de un producto en la tienda indicada.
    ⚠️ Puede tardar entre 10 y 60 segundos según la tienda.
    """
    if datos.tienda not in TIENDAS_CONFIG:
        raise HTTPException(
            status_code=400,
            detail=f"Tienda no válida. Opciones: {list(TIENDAS_CONFIG.keys())}"
        )

    try:
        url, productos = hacer_scraping(datos.producto, datos.tienda)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error de scraping: {str(e)}")

    if not productos:
        raise HTTPException(status_code=404, detail="No se encontraron productos")

    # Guardar búsqueda
    busqueda = models.Busqueda(
        usuario_id = usuario.id,
        producto   = datos.producto,
        tienda     = datos.tienda,
        url        = url,
        total      = len(productos),
    )
    db.add(busqueda)
    db.flush()

    # Guardar productos
    for p in productos:
        db.add(models.Producto(
            busqueda_id  = busqueda.id,
            tienda       = p["tienda"],
            titulo       = p["titulo"],
            precio_texto = p["precio_texto"],
            precio       = p["precio"],
        ))

    db.commit()
    db.refresh(busqueda)
    return busqueda


@app.get(
    "/scraper/historial",
    response_model=list[schemas.BusquedaResumen],
    tags=["🤖 Scraper"],
    summary="Ver historial de búsquedas",
)
def historial(
    tienda:   Optional[str] = Query(None, description="Filtrar por tienda"),
    producto: Optional[str] = Query(None, description="Filtrar por nombre de producto"),
    limite:   int           = Query(20, ge=1, le=100, description="Máximo de resultados"),
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """Devuelve el historial de búsquedas del usuario autenticado."""
    query = db.query(models.Busqueda).filter(models.Busqueda.usuario_id == usuario.id)

    if tienda:
        query = query.filter(models.Busqueda.tienda.ilike(f"%{tienda}%"))
    if producto:
        query = query.filter(models.Busqueda.producto.ilike(f"%{producto}%"))

    return query.order_by(models.Busqueda.creada_en.desc()).limit(limite).all()


@app.get(
    "/scraper/busqueda/{busqueda_id}",
    response_model=schemas.BusquedaRespuesta,
    tags=["🤖 Scraper"],
    summary="Ver detalle de una búsqueda",
)
def detalle_busqueda(
    busqueda_id: int,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """Devuelve todos los productos de una búsqueda específica."""
    busqueda = db.query(models.Busqueda).filter(
        models.Busqueda.id == busqueda_id,
        models.Busqueda.usuario_id == usuario.id,
    ).first()

    if not busqueda:
        raise HTTPException(status_code=404, detail="Búsqueda no encontrada")
    return busqueda


@app.delete(
    "/scraper/busqueda/{busqueda_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    tags=["🤖 Scraper"],
    summary="Eliminar una búsqueda",
)
def eliminar_busqueda(
    busqueda_id: int,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """Elimina una búsqueda y todos sus productos del historial."""
    busqueda = db.query(models.Busqueda).filter(
        models.Busqueda.id == busqueda_id,
        models.Busqueda.usuario_id == usuario.id,
    ).first()

    if not busqueda:
        raise HTTPException(status_code=404, detail="Búsqueda no encontrada")

    db.delete(busqueda)
    db.commit()


# ══════════════════════════════════════════════════════════════════════════════
#  COMPARADOR
# ══════════════════════════════════════════════════════════════════════════════

@app.get(
    "/comparador/comparar",
    response_model=schemas.ComparacionRespuesta,
    tags=["📊 Comparador"],
    summary="Comparar precios de un producto entre tiendas",
)
def comparar(
    producto: str = Query(..., description="Nombre del producto a comparar", examples=["laptop"]),
    top_n:    int = Query(5, ge=1, le=20, description="Cantidad de productos en el top"),
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """
    Compara precios del producto buscado en todas las tiendas del historial del usuario.
    Agrupa los resultados, calcula estadísticas y devuelve el top de más baratos y más caros.
    """
    # Obtener todas las búsquedas del usuario que coincidan con el producto
    busquedas = db.query(models.Busqueda).filter(
        models.Busqueda.usuario_id == usuario.id,
        models.Busqueda.producto.ilike(f"%{producto}%"),
    ).all()

    if not busquedas:
        raise HTTPException(
            status_code=404,
            detail=f"No hay búsquedas de '{producto}' en tu historial. Busca primero en /scraper/buscar"
        )

    # Recopilar todos los productos
    todos = []
    for b in busquedas:
        todos.extend(b.productos)

    if not todos:
        raise HTTPException(status_code=404, detail="Las búsquedas no tienen productos")

    # Estadísticas por tienda
    por_tienda: dict[str, list[float]] = {}
    for p in todos:
        if p.precio and p.precio > 0:
            por_tienda.setdefault(p.tienda, []).append(p.precio)

    estadisticas = []
    for tienda, precios in por_tienda.items():
        estadisticas.append(schemas.EstadisticaTienda(
            tienda          = tienda,
            total_productos = len(precios),
            precio_promedio = round(sum(precios) / len(precios), 2),
            precio_mediana  = round(statistics.median(precios), 2),
            precio_minimo   = round(min(precios), 2),
            precio_maximo   = round(max(precios), 2),
        ))

    estadisticas.sort(key=lambda x: x.precio_promedio or float("inf"))

    tienda_mas_barata = estadisticas[0].tienda  if estadisticas else None
    tienda_mas_cara   = estadisticas[-1].tienda if len(estadisticas) > 1 else None
    diferencia_pct    = None
    if tienda_mas_barata and tienda_mas_cara and tienda_mas_barata != tienda_mas_cara:
        prom_barata = estadisticas[0].precio_promedio
        prom_cara   = estadisticas[-1].precio_promedio
        if prom_barata and prom_barata > 0:
            diferencia_pct = round(((prom_cara - prom_barata) / prom_barata) * 100, 1)

    # Top más baratos y más caros (con precio válido)
    con_precio = [p for p in todos if p.precio and p.precio > 0]
    top_baratos = sorted(con_precio, key=lambda x: x.precio)[:top_n]
    top_caros   = sorted(con_precio, key=lambda x: x.precio, reverse=True)[:top_n]

    return schemas.ComparacionRespuesta(
        producto           = producto,
        tiendas_comparadas = len(estadisticas),
        estadisticas       = estadisticas,
        tienda_mas_barata  = tienda_mas_barata,
        tienda_mas_cara    = tienda_mas_cara,
        diferencia_pct     = diferencia_pct,
        top_baratos        = top_baratos,
        top_caros          = top_caros,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTAR
# ══════════════════════════════════════════════════════════════════════════════

def _obtener_productos_df(busqueda_id: int, usuario_id: int, db: Session) -> pd.DataFrame:
    """Obtiene los productos de una búsqueda como DataFrame."""
    busqueda = db.query(models.Busqueda).filter(
        models.Busqueda.id == busqueda_id,
        models.Busqueda.usuario_id == usuario_id,
    ).first()

    if not busqueda:
        raise HTTPException(status_code=404, detail="Búsqueda no encontrada")

    datos = [
        {
            "Tienda":         p.tienda,
            "Título":         p.titulo,
            "Precio (texto)": p.precio_texto,
            "Precio (MXN)":   p.precio,
            "Fecha":          p.creado_en.strftime("%d/%m/%Y %H:%M"),
        }
        for p in busqueda.productos
    ]
    return pd.DataFrame(datos), busqueda


@app.get(
    "/exportar/csv/{busqueda_id}",
    tags=["📥 Exportar"],
    summary="Descargar resultados en CSV",
)
def exportar_csv(
    busqueda_id: int,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """Descarga los productos de una búsqueda como archivo CSV."""
    df, busqueda = _obtener_productos_df(busqueda_id, usuario.id, db)

    buffer = io.StringIO()
    df.to_csv(buffer, index=False, encoding="utf-8-sig")
    buffer.seek(0)

    nombre = f"precios_{busqueda.tienda.replace(' ', '_').lower()}_{busqueda.producto.replace(' ', '_')}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@app.get(
    "/exportar/excel/{busqueda_id}",
    tags=["📥 Exportar"],
    summary="Descargar resultados en Excel",
)
def exportar_excel(
    busqueda_id: int,
    db: Session = Depends(get_db),
    usuario: models.Usuario = Depends(get_usuario_actual),
):
    """Descarga los productos de una búsqueda como archivo Excel formateado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    df, busqueda = _obtener_productos_df(busqueda_id, usuario.id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    encabezados = ["#", "Tienda", "Título", "Precio (texto)", "Precio (MXN)"]
    for ci, h in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = PatternFill("solid", start_color="2F3640")
        c.alignment = Alignment(horizontal="center")

    # Datos
    fill_par = PatternFill("solid", start_color="F4F6FA")
    for i, row in enumerate(df.itertuples(index=False), 1):
        fill = fill_par if i % 2 == 0 else PatternFill()
        ws.cell(row=i+1, column=1, value=i).fill = fill
        ws.cell(row=i+1, column=2, value=row[0]).fill = fill
        c_tit = ws.cell(row=i+1, column=3, value=row[1])
        c_tit.alignment = Alignment(wrap_text=True)
        c_tit.fill = fill
        ws.cell(row=i+1, column=4, value=row[2]).fill = fill
        c_pre = ws.cell(row=i+1, column=5, value=row[3])
        c_pre.number_format = "$#,##0.00"
        c_pre.fill = fill

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 15
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"precios_{busqueda.tienda.replace(' ', '_').lower()}_{busqueda.producto.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ── Health check ───────────────────────────────────────────────────────────────

@app.get("/", tags=["⚙️ Sistema"], summary="Health check")
def root():
    return {
        "status":  "ok",
        "mensaje": "API Scraper de Precios funcionando 🚀",
        "docs":    "http://localhost:8000/docs",
    }